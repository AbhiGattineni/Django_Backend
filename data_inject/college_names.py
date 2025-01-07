import pandas as pd
import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook


def get_hyperlinks(sheet, index):
    hyperlinks = {}
    temp = 0
    for row in sheet.iter_rows(min_row=2):
        if temp == index:
            for cell in row:
                if cell.hyperlink:
                    hyperlinks[cell.coordinate] = cell.hyperlink.target
                else:
                    hyperlinks[cell.coordinate] = None
        temp=temp+1
    return hyperlinks

def insert_data_to_mysql(df, sheetname, cursor, conn):
    for index, row in df.iterrows():
        print("row",index)
        wb = load_workbook(excel_file, data_only=False)
        sheet = wb[sheetname]
        hyperlinks = get_hyperlinks(sheet,index)
        if pd.isna(row['COLLEGE NAME']):
            break
        sql = """
        INSERT INTO collegelist (
            college_name, website_link, international_UG_link, international_graduation_link, 
            application_UG_link, application_graduation_link, application_UG_fee, application_UG_fee_link, 
            application_graduation_fee, application_graduation_fee_link, gre_score, gre_score_link, 
            toefl_UG_score, toefl_UG_score_link, toefl_graduation_score, toefl_graduation_score_link, 
            ielts_ug_score, ielts_ug_score_link, ielts_graduation_score, ielts_graduation_score_link, 
            fall_deadline_UG, fall_deadline_UG_link, fall_deadline_graduation, fall_deadline_graduation_link, 
            spring_deadline_UG, spring_deadline_UG_link, spring_deadline_graduation, spring_deadline_graduation_link, 
            college_email, college_email_link, college_phone, college_phone_link, 
            international_person_email, international_person_email_link, public_private, 
            UG_courses, UG_courses_link, graduation_courses, graduation_courses_link, state
        ) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        values = (
            row['COLLEGE NAME'], 
            row['WEBSITE LINK'] if not pd.isna(row['WEBSITE LINK']) else None, 
            row['International Section Link UG'] if not pd.isna(row['International Section Link UG']) else None, 
            row['International Section Link Graduate'] if not pd.isna(row['International Section Link Graduate']) else None, 
            row['Application Link UG'] if not pd.isna(row['Application Link UG']) else None, 
            row['Application Link Graduate'] if not pd.isna(row['Application Link Graduate']) else None, 
            row['Application Fee UG'] if not pd.isna(row['Application Fee UG']) else None,
            hyperlinks[f'H{index+2}'],  # Placeholder for application UG fee link
            row['Application Fee Graduate'] if not pd.isna(row['Application Fee Graduate']) else None,
            hyperlinks[f'I{index+2}'],  # Placeholder for application graduation fee link
            row['GRE SCORE'] if not pd.isna(row['GRE SCORE']) else None,
            hyperlinks[f'J{index+2}'],  # Placeholder for GRE score link
            row['TOEFL SCORE UG'] if not pd.isna(row['TOEFL SCORE UG']) else None,
            hyperlinks[f'M{index+2}'],  # Placeholder for TOEFL score link
            row['TOEFL SCORE Graduate'] if not pd.isna(row['TOEFL SCORE Graduate']) else None,
            hyperlinks[f'N{index+2}'],  # Placeholder for TOEFL score link
            row['IELTS SCORE UG'] if not pd.isna(row['IELTS SCORE UG']) else None,
            hyperlinks[f'K{index+2}'],  # Placeholder for TOEFL score link
            row['IELTS SCORE Graduate'] if not pd.isna(row['IELTS SCORE Graduate']) else None,
            hyperlinks[f'L{index+2}'],  # Placeholder for TOEFL score link
            row['FALL DEADLINE UG'] if not pd.isna(row['FALL DEADLINE UG']) else None,
            hyperlinks[f'O{index+2}'],  # Placeholder for fall deadline link
            row['FALL DEADLINE Graduate'] if not pd.isna(row['FALL DEADLINE Graduate']) else None,
            hyperlinks[f'P{index+2}'],  # Placeholder for fall deadline link
            row['SPRING DEADLINE UG'] if not pd.isna(row['SPRING DEADLINE UG']) else None,
            hyperlinks[f'Q{index+2}'],  # Placeholder for spring deadline link
            row['SPRING DEADLINE Graduate'] if not pd.isna(row['SPRING DEADLINE Graduate']) else None,
            hyperlinks[f'R{index+2}'],  # Placeholder for spring deadline link
            row['COLLEGE EMAIL ID'] if not pd.isna(row['COLLEGE EMAIL ID']) else None,
            hyperlinks[f'S{index+2}'],  # Placeholder for email link
            row['PHONE NUMBER'] if not pd.isna(row['PHONE NUMBER']) else None,
            hyperlinks[f'T{index+2}'],  # Placeholder for phone link
            row['PERSON EMAID ID'] if not pd.isna(row['PERSON EMAID ID']) else None,
            hyperlinks[f'U{index+2}'],  # Placeholder for email link
            row['PUBLIC/PRIVATE']  if not pd.isna(row['PUBLIC/PRIVATE']) else None, 
            row['COURSES AVAILABLE UG'] if not pd.isna(row['COURSES AVAILABLE UG']) else None,
            hyperlinks[f'W{index+2}'],  # Placeholder for UG courses link
            row['COURSES AVAILABLE Graduate'] if not pd.isna(row['COURSES AVAILABLE Graduate']) else None,
            hyperlinks[f'X{index+2}'],  # Placeholder for graduation courses link
            sheetname  # Placeholder for state (if needed)
        )

        # Debugging output
        # print(f"SQL: {sql}")
        # print(f"Values: {values}")
        # print(f"Number of placeholders: {sql.count('%s')}, Number of values: {len(values)}")

        try:
            cursor.execute(sql, values)
            conn.commit()
        except Error as e:
            print(f"Error: {e}")
            conn.rollback()

def read_excel_and_insert_to_mysql(excel_file, db_config):
    # Read the Excel file with multiple sheets
    xls = pd.ExcelFile(excel_file)
    
    # Connect to MySQL database
    try:
        conn = mysql.connector.connect(**db_config)
        if conn.is_connected():
            print("Connected to MySQL")
            cursor = conn.cursor()

            # Iterate through the sheets and insert data into corresponding MySQL tables
            for sheet_name in xls.sheet_names:
                if sheet_name == 'Sheet49':
                    # Load the sheet into a DataFrame
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    colleges = df['COLLEGE NAME'].tolist()
                    data = {
                        "id" : range(1081,1081+len(colleges)),
                        "college names" : colleges
                    }
                    new_df = pd.DataFrame(data)
                    new_df.to_excel(r"C:\Users\thrived\Downloads\USA Colleges updated.xlsx",index=False)
                    # Insert data into a MySQL table
                    insert_data_to_mysql(df, sheet_name, cursor, conn)
            print("Data insertion completed.")
    except Error as e:
        print(f"Error while connecting to MySQL: {e}")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            print("MySQL connection is closed")

# Database configuration
db_config = {
    'host': 'database.ctnj0eswvoik.us-west-2.rds.amazonaws.com',
    'user': 'admin',
    'password': 'AbhishekAWS1234',
    'database': 'MYBLOG'
}

# Specify the path to your Excel file
excel_file = r"C:\Users\thrived\Downloads\USA Colleges new.xlsx"

# Call the function to read Excel and insert data into MySQL
read_excel_and_insert_to_mysql(excel_file, db_config)
