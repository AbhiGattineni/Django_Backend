from django.core.management.base import BaseCommand
import pandas as pd
from backend_api.models import CollegesList
from openpyxl import load_workbook
from math import isnan
from decimal import Decimal, InvalidOperation
from django.utils.dateparse import parse_date



class Command(BaseCommand):
    help = 'Import data from Excel sheet into the database'

    def get_hyperlink(self, excel_sheet, row_num, column_name, df):
        # Find the column index (1-based) for the given column name
        column_index = df.columns.get_loc(column_name) + 1

        # Get the cell at the specified row and column
        cell = excel_sheet.cell(row=row_num, column=column_index)
        return cell.hyperlink.target if cell.hyperlink else None
    
    def get_decimal_value(self, value):
        try:
            # Check if value is 'nan'
            if isinstance(value, float) and isnan(value):
                return None
            # Convert to float first to handle strings like '1.23'
            return Decimal(str(float(value)))
        except (ValueError, TypeError, InvalidOperation):
            return None
    
    def get_value_as_string(self, value):
        if value is None or pd.isna(value):
            return None  
        return str(value)


    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to the Excel file')
        parser.add_argument('tabs', nargs='+', type=str, help='Names of tabs to import data from')

    def handle(self, *args, **options):
        file_path = options['file_path']
        tabs = options['tabs']

        try:
            wb = load_workbook(file_path, data_only=True)
            with pd.ExcelFile(file_path) as xls:
                for tab in tabs:
                    sheet = wb[tab]
                    # Read the Excel sheet for the current tab
                    df = pd.read_excel(xls, sheet_name=tab, engine='openpyxl', header=0)


                    # Loop through rows and create College objects
                    for index, row in df.iterrows():
                        row_num = index + 2  # Adjust based on header row and zero-indexing
                
                        college = CollegesList(
                            college_name=row['COLLEGE NAME'],
                            website_link=row['WEBSITE LINK'],
                            international_UG_link=row['International Section Link UG'],
                            international_graduation_link=row['International Section Link Graduate'],  
                            application_UG_link=row['Application Link UG'],  
                            application_graduation_link=row['Application Link Graduate'],  
                            application_UG_fee= row['Application Fee UG'],  
                            application_UG_fee_link=self.get_hyperlink(sheet, row_num, 'Application Fee UG',df),
                            application_graduation_fee=row['Application Fee Graduate'] if not pd.isna(row['Application Fee Graduate']) else None,
                            application_graduation_fee_link=self.get_hyperlink(sheet, row_num, 'Application Fee Graduate',df),  
                            gre_score= row['GRE SCORE'],   
                            gre_score_link=self.get_hyperlink(sheet, row_num, 'GRE SCORE',df),  
                            toefl_UG_score= row['TOEFL SCORE UG'], 
                            toefl_UG_score_link=self.get_hyperlink(sheet, row_num, 'TOEFL SCORE UG',df),  
                            toefl_graduation_score= row['TOEFL SCORE Graduate'], 
                            toefl_graduation_score_link=self.get_hyperlink(sheet, row_num, 'TOEFL SCORE Graduate',df),  
                            ielts_ug_score=row['IELTS SCORE UG'],  
                            ielts_ug_score_link=self.get_hyperlink(sheet, row_num, 'IELTS SCORE UG',df),  
                            ielts_graduation_score=self.get_value_as_string(row['IELTS SCORE Graduate']),  
                            ielts_graduation_score_link=self.get_hyperlink(sheet, row_num, 'IELTS SCORE Graduate',df), 
                            fall_deadline_UG=str(row['FALL DEADLINE UG']),  
                            fall_deadline_UG_link=self.get_hyperlink(sheet, row_num, 'FALL DEADLINE UG',df),   
                            fall_deadline_graduation=str(row['FALL DEADLINE Graduate']),  
                            fall_deadline_graduation_link=self.get_hyperlink(sheet, row_num, 'FALL DEADLINE Graduate',df),
                            spring_deadline_UG=row['SPRING DEADLINE UG'],
                            spring_deadline_UG_link=self.get_hyperlink(sheet, row_num, 'SPRING DEADLINE UG',df),
                            spring_deadline_graduation=str(row['SPRING DEADLINE Graduate']),  
                            spring_deadline_graduation_link=self.get_hyperlink(sheet, row_num, 'SPRING DEADLINE Graduate',df),
                            college_email=row['COLLEGE EMAIL ID'],  
                            college_email_link=self.get_hyperlink(sheet, row_num, 'COLLEGE EMAIL ID',df),  
                            college_phone=row['PHONE NUMBER'],  
                            college_phone_link=self.get_hyperlink(sheet, row_num, 'PHONE NUMBER',df),
                            international_person_email=row['PERSON EMAID ID'],  
                            international_person_email_link=self.get_hyperlink(sheet, row_num, 'PERSON EMAID ID',df),
                            public_private=row['PUBLIC/PRIVATE'],  
                            UG_courses=row['COURSES AVAILABLE UG'],  
                            UG_courses_link=self.get_hyperlink(sheet, row_num, 'COURSES AVAILABLE UG',df),
                            graduation_courses=row['COURSES AVAILABLE Graduate'],  
                            )
                        college.save()

                    self.stdout.write(self.style.SUCCESS(f'Data from tab "{tab}" imported successfully'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'An error occurred: {e}'))